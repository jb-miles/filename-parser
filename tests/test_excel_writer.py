#!/usr/bin/env python3
"""
Tests for Excel writer formatting helpers.
"""

from __future__ import annotations

from openpyxl import load_workbook

from modules.excel_writer import ExcelSheetData, write_excel_workbook


def test_excel_writer_bolds_passthrough_cells(tmp_path):
    output_path = tmp_path / "out.xlsx"
    sheet = ExcelSheetData(
        name="Test",
        headers=["a", "b"],
        rows=[[1, 2], [3, 4]],
        bold_cells=[[True, False], [False, True]],
    )

    write_excel_workbook(output_path, [sheet])

    wb = load_workbook(output_path)
    try:
        ws = wb["Test"]
        assert ws.cell(row=2, column=1).font.bold is True
        assert ws.cell(row=2, column=2).font.bold is not True
        assert ws.cell(row=3, column=1).font.bold is not True
        assert ws.cell(row=3, column=2).font.bold is True
    finally:
        wb.close()

