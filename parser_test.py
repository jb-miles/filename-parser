#!/usr/bin/env python3
"""
Test script for filename parser.
Reads filenames from a text file and outputs results to Excel.
"""

import csv
import argparse
import sys
import json
import os
try:
    from .parser import FilenameParser
    from .modules import PreTokenizer
except ImportError:
    from parser import FilenameParser
    from modules import PreTokenizer
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import re


def is_fully_identified_pattern(pattern_str):
    """
    Check if a pattern contains only labeled tokens (no generic tokenN).
    A pattern is fully identified if it has no references to 'tokenN' where N is a number.
    """
    if not pattern_str:
        return False
    # Look for patterns like token0, token1, token2, etc.
    return not re.search(r'\btoken\d+\b', pattern_str)


def main():
    """Process filenames from input file and write results to Excel."""

    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Process filenames and output results to Excel')
    parser.add_argument('input_file', nargs='?', help='Input file containing filenames (one per line)')
    parser.add_argument('output_file', nargs='?', help='Output Excel file')
    parser.add_argument('-p', '--pre-tokenize-only', action='store_true',
                       help='Only perform pre-tokenization and output original, removed, and cleaned')
    
    args = parser.parse_args()
    
    # Configuration
    if args.input_file:
        input_file = args.input_file
        # Default to .xlsx if no extension provided
        if args.output_file:
            output_file = args.output_file
            if not output_file.endswith('.xlsx'):
                output_file = output_file.replace('.csv', '.xlsx') if output_file.endswith('.csv') else output_file + '.xlsx'
        else:
            output_file = input_file.replace('.txt', '-results.xlsx')
    else:
        input_file = "/Users/jbmiles/Documents/sample.txt"
        output_file = "/Users/jbmiles/Documents/sample-results.xlsx"

    filename_parser = FilenameParser()
    pre_tokenizer = PreTokenizer()

    print(f"Reading from: {input_file}")
    print(f"Writing to: {output_file}")
    if args.pre_tokenize_only:
        print("Mode: Pre-tokenization only")

    # Process filenames
    # Try to detect encoding and handle various input encodings
    import codecs
    import unicodedata
    
    def open_file_with_encoding_detection(filepath, mode='r'):
        """Open file with encoding detection for common encodings."""
        encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1']
        
        for encoding in encodings:
            try:
                return open(filepath, mode, encoding=encoding)
            except UnicodeDecodeError:
                continue
        
        # If all fail, try with errors='replace'
        return open(filepath, mode, encoding='utf-8', errors='replace')
    
    def normalize_unicode_text(text):
        """Normalize Unicode text to handle different character representations."""
        # Normalize to NFC form (canonical decomposition followed by composition)
        # This helps handle cases where the same character might be represented differently
        normalized = unicodedata.normalize('NFC', text)
        
        # Replace common problematic characters
        replacements = {
            '\u201c': '"',  # Left double quotation mark
            '\u201d': '"',  # Right double quotation mark
            '\u2018': "'",  # Left single quotation mark
            '\u2019': "'",  # Right single quotation mark
            '\u2013': '-',  # En dash
            '\u2014': '--', # Em dash
            '\u2026': '...', # Horizontal ellipsis
            '\u00a0': ' ',  # Non-breaking space
            '\u200b': '',   # Zero-width space
            '\u200e': '',   # Left-to-right mark
            '\u200f': '',   # Right-to-left mark
        }
        
        for old, new in replacements.items():
            normalized = normalized.replace(old, new)
            
        return normalized
    
    # Store existing column widths and table info if file exists
    existing_column_widths = {}
    existing_table_info = None
    if os.path.exists(output_file):
        try:
            existing_wb = load_workbook(output_file)
            if existing_wb.worksheets:
                existing_ws = existing_wb.active
                if existing_ws:
                    # Get column dimensions
                    for col_letter, col_dim in existing_ws.column_dimensions.items():
                        if col_dim.width:
                            existing_column_widths[col_letter] = col_dim.width

                    # Check for existing table
                    if existing_ws.tables:
                        # Get the first table's info (usually only one table per sheet)
                        first_table_name = list(existing_ws.tables.keys())[0]
                        existing_table = existing_ws.tables[first_table_name]
                        existing_table_info = {
                            'name': first_table_name,
                            'style': existing_table.tableStyleInfo.name if existing_table.tableStyleInfo else None,
                            'has_filters': existing_table.tableStyleInfo.showRowStripes if existing_table.tableStyleInfo else True
                        }
            existing_wb.close()
        except Exception as e:
            print(f"Warning: Could not read existing file properties: {e}")

    # Create new workbook
    wb = Workbook()
    ws = wb.active
    if ws:
        ws.title = "Filename Parser Results"
    
    # Define cell styles
    date_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue
    studio_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    studio_code_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Light yellow
    performers_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red

    # Styles for fully-identified patterns
    fully_identified_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Bright green
    fully_identified_font = Font(bold=True)

    # We'll determine the header after processing all filenames to find max tokens
    header_written = False
    header = []
    json_keys = []
    all_rows_data = []  # Store all row data to determine max tokens
    token_types = {}  # Store token types for coloring
    
    line_count = 0
    with open_file_with_encoding_detection(input_file) as f_in:
        for line in f_in:
            line = line.strip()
            
            # Normalize Unicode characters
            line = normalize_unicode_text(line)

            # Skip empty lines and section markers
            if not line or line in ['exceptions', 'sacrifice']:
                continue

            # Process the filename
            if args.pre_tokenize_only:
                result = pre_tokenizer.process(line)
                json_output = result.to_json()
                json_data = json.loads(json_output)
                
                # If this is the first row, write the header based on JSON keys
                if not header_written:
                    # Use keys from the JSON output plus Original, Cleaned
                    json_keys = list(json_data.keys())
                    header = ['Original'] + json_keys + ['Cleaned']
                    # Write header to Excel
                    if ws:
                        for col_idx, col_name in enumerate(header, 1):
                            ws.cell(row=1, column=col_idx, value=col_name)
                    header_written = True
                
                # Prepare the row data
                row_data = {'Original': result.original}
                
                # Add values for each JSON key
                for key in json_keys:
                    if key == 'removed_tokens':
                        # Special handling for removed_tokens array
                        tokens_str = ' | '.join([f"{t['value']}({t['category']})" for t in json_data[key]])
                        row_data[key] = tokens_str
                    else:
                        row_data[key] = str(json_data[key])
                
                # Add the cleaned data
                row_data['Cleaned'] = result.cleaned
                
                # Write the row in the correct column order
                if ws:
                    for col_idx, col_name in enumerate(header, 1):
                        ws.cell(row=line_count + 2, column=col_idx, value=row_data.get(col_name, ''))
            else:
                # Full processing pipeline - use parse() to run all steps
                final_result = filename_parser.parse(line)
                json_output = final_result.to_json()
                json_data = json.loads(json_output)

                # Get removed tokens separately for the "Removed" column
                pre_result = filename_parser.pre_tokenize(line)
                removed_str = ' | '.join([f"{t.value}({t.category})" for t in pre_result.removed_tokens])

                # Store row data for later processing
                row_data = {
                    'Original': final_result.original,
                    'Removed': removed_str,
                    'Cleaned': final_result.cleaned,
                    'path': json_data.get('path', ''),
                    'pattern': json_data.get('pattern', ''),
                    **json_data  # Include all token fields
                }
                
                # Store token types for coloring
                row_token_types = {}
                if final_result and final_result.tokens:
                    for token in final_result.tokens:
                        if token.type in ['date', 'studio', 'studio_code', 'performers']:
                            # Find which column this token would be in
                            token_key = None
                            for key in json_data.keys():
                                if key.startswith('token') and json_data[key] == token.value:
                                    token_key = key
                                    break
                            if token_key:
                                row_token_types[token_key] = token.type
                
                token_types[line_count] = row_token_types
                all_rows_data.append(row_data)
            line_count += 1
        
        # After processing all rows, determine the header with max tokens
        if not header_written and all_rows_data:
            # Find all token keys across all rows
            all_token_keys = set()
            for row_data in all_rows_data:
                all_token_keys.update(key for key in row_data.keys() if key.startswith('token'))
            
            # Determine the maximum number of tokens
            max_tokens = 0
            if all_token_keys:
                token_numbers = [int(key[5:]) for key in all_token_keys if key[5:].isdigit()]
                max_tokens = max(token_numbers) + 1 if token_numbers else 0
            
            # Create header with Original, Removed, Cleaned, path, pattern, and token0 through tokenN
            header = ['Original', 'Removed', 'Cleaned', 'path', 'pattern']
            for i in range(max_tokens):
                header.append(f"token{i}")
            
            # Write header to Excel
            if ws:
                for col_idx, col_name in enumerate(header, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)
            header_written = True
            
            # Separate fully-identified rows from others
            fully_identified_rows = []
            other_rows = []

            for row_idx, row_data in enumerate(all_rows_data):
                pattern = row_data.get('pattern', '')
                if is_fully_identified_pattern(pattern):
                    fully_identified_rows.append((row_idx, row_data))
                else:
                    other_rows.append((row_idx, row_data))

            # Combine: fully-identified rows first, then others
            ordered_rows = fully_identified_rows + other_rows

            # Write all rows in the new order
            for output_row_idx, (original_row_idx, row_data) in enumerate(ordered_rows):
                # Ensure all token columns are present
                for col in header:
                    if col.startswith('token') and col not in row_data:
                        row_data[col] = ''

                # Write the row in the correct column order
                if ws:
                    pattern = row_data.get('pattern', '')
                    is_fully_identified = is_fully_identified_pattern(pattern)

                    for col_idx, col_name in enumerate(header, 1):
                        cell_value = row_data.get(col_name, '')
                        cell = ws.cell(row=output_row_idx + 2, column=col_idx, value=cell_value)

                        # Apply pattern cell styling for fully-identified rows
                        if col_name == 'pattern' and is_fully_identified:
                            cell.fill = fully_identified_fill
                            cell.font = fully_identified_font

                        # Apply coloring based on token type
                        if col_name.startswith('token') and original_row_idx in token_types:
                            token_type = token_types[original_row_idx].get(col_name)
                            if token_type == 'date':
                                cell.fill = date_fill
                            elif token_type == 'studio_code':
                                cell.fill = studio_code_fill
                            elif token_type == 'studio':
                                cell.fill = studio_fill
                            elif token_type == 'performers':
                                cell.fill = performers_fill
    
    # Restore column widths if they existed
    if ws:
        if existing_column_widths:
            for column_letter, width in existing_column_widths.items():
                if width:
                    ws.column_dimensions[column_letter].width = width
        else:
            # Auto-adjust column widths if no previous widths
            # Use a different approach to iterate through columns
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        col_letter = cell.column_letter
                        current_width = ws.column_dimensions[col_letter].width or 0

                        try:
                            cell_length = len(str(cell.value))
                            # Set minimum width and add some padding
                            adjusted_width = max(cell_length + 2, 10)

                            # Only update if this width is larger than current
                            if adjusted_width > current_width:
                                ws.column_dimensions[col_letter].width = adjusted_width
                        except:
                            pass

        # Create or update table
        if header_written and line_count > 0:
            # Calculate the data range (A1:LastColumn + LastRow)
            last_col = get_column_letter(len(header))
            data_range = f"A1:{last_col}{line_count + 1}"  # +1 for header row

            # Remove existing table if present
            if ws.tables:
                for table_name in list(ws.tables.keys()):
                    del ws.tables[table_name]

            # Create new table
            table_style = existing_table_info['style'] if existing_table_info else 'TableStyleMedium9'
            table = Table(displayName="ParseResults", ref=data_range)
            style = TableStyleInfo(name=table_style, showFirstColumn=False,
                                  showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)

        # Save the workbook
        wb.save(output_file)

    print(f"Processed {line_count} filenames")
    print(f"Results written to {output_file}")
    
    # Print all named variables
    print("\n=== Named Variables ===")
    local_vars = {k: v for k, v in locals().items() if not k.startswith('__')}
    for var_name, var_value in local_vars.items():
        if isinstance(var_value, (str, int, float, bool)):
            print(f"{var_name}: {var_value}")
        elif isinstance(var_value, list) and len(var_value) < 10:
            print(f"{var_name}: {var_value}")
        elif isinstance(var_value, dict) and len(var_value) < 10:
            print(f"{var_name}: {var_value}")
        else:
            print(f"{var_name}: {type(var_value).__name__} (length: {len(var_value) if hasattr(var_value, '__len__') else 'N/A'})")


if __name__ == '__main__':
    main()
