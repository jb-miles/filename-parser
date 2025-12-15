#!/usr/bin/env python3
"""
Utility helpers for writing Excel reports in a consistent table style.

Provides thin wrappers around openpyxl so both the evaluation harness and the
Stash plugin can share the same formatting logic (headers, auto-width, tables).
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


HighlightPredicate = Callable[[Any], bool]


@dataclass(frozen=True)
class ExcelSheetData:
    """
    Describes a sheet to be written to the workbook.

    Attributes:
        name: Sheet/tab name.
        headers: Ordered list of column headers.
        rows: Iterable of row values already ordered to match headers.
        highlight_discrepancies: Whether to apply yellow fill when predicate matches.
        discrepancy_predicate: Optional predicate used when highlighting is enabled.
    """

    name: str
    headers: Sequence[str]
    rows: Sequence[Sequence[Any]]
    highlight_discrepancies: bool = False
    discrepancy_predicate: Optional[HighlightPredicate] = None
    bold_cells: Optional[Sequence[Sequence[bool]]] = None


def _write_excel_sheet(ws, sheet: ExcelSheetData) -> None:
    """Render a single sheet using provided headers/rows and optional highlighting."""
    ws.title = sheet.name

    headers = list(sheet.headers)
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    should_highlight = sheet.highlight_discrepancies and sheet.discrepancy_predicate
    bold_font = Font(bold=True)

    for row_idx, row in enumerate(sheet.rows, 2):
        bold_row: Optional[Sequence[bool]] = None
        if sheet.bold_cells is not None and (row_idx - 2) < len(sheet.bold_cells):
            bold_row = sheet.bold_cells[row_idx - 2]

        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if should_highlight and sheet.discrepancy_predicate(value):
                cell.fill = yellow_fill
            if bold_row is not None and (col_idx - 1) < len(bold_row) and bool(bold_row[col_idx - 1]):
                cell.font = bold_font

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(headers[col_idx - 1])

        for cell in ws[col_letter]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    if sheet.rows:
        last_col = get_column_letter(len(headers))
        data_range = f"A1:{last_col}{len(sheet.rows) + 1}"
        table_name = sheet.name.replace(" ", "") + "Table"
        table = Table(displayName=table_name, ref=data_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)


def write_excel_workbook(output_path: Path | str, sheets: Sequence[ExcelSheetData]) -> Path:
    """
    Write a workbook consisting of the provided sheets.

    Args:
        output_path: Destination path for the workbook.
        sheets: Ordered sheet definitions to render.

    Returns:
        Path to the written workbook.
    """
    if not sheets:
        raise ValueError("At least one sheet must be provided to write a workbook.")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    for idx, sheet in enumerate(sheets):
        ws = wb.active if idx == 0 else wb.create_sheet()
        _write_excel_sheet(ws, sheet)

    wb.save(output_path)
    return output_path
