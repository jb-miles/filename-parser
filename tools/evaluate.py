#!/usr/bin/env python3
"""
Evaluation harness for filename parser.

Provides two modes:
- blind: Coverage-first metrics without reference labels
- reference: Accuracy metrics vs expected labels

Outputs:
- Excel workbook with parsed results
- JSON metrics file for automation
"""

import sys
import argparse
import json
import ast
from datetime import datetime
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Union
from collections import Counter

# Add parent directory to path to import parser modules
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from yansa import FilenameParser
from modules import PreTokenizer
from openpyxl import load_workbook

from modules.excel_writer import ExcelSheetData, write_excel_workbook


@dataclass
class ParsedRow:
    """Represents a single parsed filename with all extracted fields."""
    input: str
    removed: str
    # path: Optional[str]  # Disabled - not working on paths yet
    filename_cleaned: str
    # path_pattern: Optional[str]  # Disabled - not working on paths yet
    filename_pattern: str
    studio: Optional[str]
    title: Optional[str]
    performers: Optional[str]
    date: Optional[str]
    studio_code: Optional[str]
    sequence: Optional[Dict[str, Any]]
    group: Optional[str]
    # unlabeled_path_tokens: Optional[Set[str]]  # Disabled - not working on paths yet
    unlabeled_filename_tokens: Optional[Set[str]]
    match_stats: Dict[str, Any]

    def to_excel_row(self) -> List[Any]:
        """Convert to Excel row maintaining column order."""
        return [
            self.input,
            self.removed,
            # self.path if self.path else "",  # Disabled - not working on paths yet
            self.filename_cleaned,
            # self.path_pattern if self.path_pattern else "",  # Disabled - not working on paths yet
            self.filename_pattern,
            self.studio if self.studio else "",
            self.title if self.title else "",
            self.performers if self.performers else "",
            self.date if self.date else "",
            self.studio_code if self.studio_code else "",
            json.dumps(self.sequence) if self.sequence else "",
            self.group if self.group else "",
            # str(self.unlabeled_path_tokens) if self.unlabeled_path_tokens else "",  # Disabled - not working on paths yet
            str(self.unlabeled_filename_tokens) if self.unlabeled_filename_tokens else "",
            json.dumps(self.match_stats),
        ]

    @staticmethod
    def get_headers() -> List[str]:
        """Get Excel column headers in proper order."""
        return [
            "input",
            "removed",
            # "path",  # Disabled - not working on paths yet
            "filename_cleaned",
            # "path_pattern",  # Disabled - not working on paths yet
            "filename_pattern",
            "studio",
            "title",
            "performers",
            "date",
            "studio_code",
            "sequence",
            "group",
            # "unlabeled_path_tokens",  # Disabled - not working on paths yet
            "unlabeled_filename_tokens",
            "match_stats",
        ]


def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description='Evaluate filename parser with coverage or reference metrics'
    )
    parser.add_argument(
        '--mode',
        choices=['blind', 'reference'],
        default='blind',
        help='Evaluation mode: blind (coverage) or reference (vs labels)'
    )
    parser.add_argument(
        '--input',
        required=True,
        help='Input file containing filenames (one per line) or Excel reference'
    )
    parser.add_argument(
        '--baseline',
        help='Baseline metrics JSON for comparison (reference mode)'
    )
    parser.add_argument(
        '--output-excel',
        help='Output Excel file path (default: metrics/MODE-YYYYMMDD-HHMMSS.xlsx)'
    )
    parser.add_argument(
        '--output-json',
        help='Output JSON metrics file (default: metrics/MODE-YYYYMMDD-HHMMSS.json)'
    )
    parser.add_argument(
        '--limit',
        type=int,
        help='Limit number of files to process'
    )
    parser.add_argument(
        '--samples',
        type=int,
        default=10,
        help='Number of mismatch samples to capture (reference mode)'
    )
    parser.add_argument(
        '--no-write',
        action='store_true',
        help='Dry-run mode: skip writing output files'
    )
    parser.add_argument(
        '--skip-excel',
        action='store_true',
        help='Skip Excel output for faster CI runs'
    )

    return parser.parse_args()


def read_input_file(filepath: Union[str, Path], limit: Optional[int] = None,
                   sheet_name: Optional[str] = None) -> List[str]:
    """
    Read input file and return list of filenames.
    Handles text files and Excel files.
    """
    filepath = Path(filepath)
    filenames = []

    # Check if it's an Excel file
    if filepath.suffix == '.xlsx':
        wb = load_workbook(filepath, read_only=True)
        try:
            ws = None
            if sheet_name:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    raise ValueError(f"Could not find '{sheet_name}' sheet in Excel file. Sheets present: {wb.sheetnames}")
            else:
                ws = wb.active

            if ws is None:
                raise ValueError("Excel file has no usable worksheet")

            # Find the 'input' column (assuming first row is header)
            headers = [cell.value for cell in ws[1]]
            input_col_idx = None
            for idx, header in enumerate(headers, 1):
                normalized_header = str(header).strip().lower() if header is not None else ""
                if normalized_header == 'input':
                    input_col_idx = idx
                    break

            if input_col_idx is None:
                raise ValueError("Could not find 'input' column in Excel file")

            # Read filenames from the input column
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[input_col_idx - 1]:
                    filenames.append(str(row[input_col_idx - 1]))
                    if limit and len(filenames) >= limit:
                        break
        finally:
            wb.close()
    else:
        # Text file - read line by line
        with filepath.open('r', encoding='utf-8', errors='replace') as f:
            for line in f:
                line = line.strip()
                if line and line not in ['exceptions', 'sacrifice']:
                    filenames.append(line)
                    if limit and len(filenames) >= limit:
                        break

    return filenames


def parse_filename(parser: FilenameParser, filename: str) -> ParsedRow:
    """
    Parse a single filename and return a ParsedRow.

    This transforms the parser's output into the 13-column schema (path columns disabled).
    The parser now handles all extraction logic (title, sequence, group, etc.).
    """
    # Run full parsing pipeline
    result = parser.parse(filename)
    tokens = result.tokens or []

    # Get pre-tokenization result for removed tokens
    pre_result = parser.pre_tokenize(filename)
    removed_str = ' | '.join([f"{t.value}({t.category})" for t in pre_result.removed_tokens])

    # PATH PROCESSING DISABLED - Not working on paths yet
    # # Extract path and non-path tokens
    # path_token = None
    # filename_tokens = []
    # for token in tokens:
    #     if token.type == 'path':
    #         path_token = token
    #     else:
    #         filename_tokens.append(token)

    # All tokens are filename tokens (no path tokens)
    filename_tokens = tokens

    # PATH PROCESSING DISABLED - Not working on paths yet
    # # Build patterns
    # path_pattern = None
    # if path_token:
    #     # TODO: Build path pattern from path tokens
    #     path_pattern = None  # Placeholder for now

    # Use the pattern from the tokenizer (it includes delimiters)
    filename_pattern = result.pattern or ""

    # Extract labeled fields from parser result
    studio = result.studio
    title = result.title
    sequence = result.sequence
    group = result.group

    # Collect performers and dates from tokens
    performers = []
    date = None
    studio_code = None

    for token in tokens:
        if token.type == 'performers':
            performers.append(token.value)
        elif token.type == 'date':
            date = token.value
        elif token.type == 'studio_code':
            studio_code = token.value

    # PATH PROCESSING DISABLED - Not working on paths yet
    # # Calculate unlabeled tokens
    # unlabeled_path_tokens = set()
    # unlabeled_filename_tokens = set()
    #
    # labeled_types = {'path', 'date', 'studio', 'studio_code', 'performers', 'sequence', 'title'}
    # for token in filename_tokens:
    #     if token.type not in labeled_types and token.value.strip():
    #         unlabeled_filename_tokens.add(token.value)

    # Calculate unlabeled filename tokens only
    unlabeled_filename_tokens = set()
    labeled_types = {'date', 'studio', 'studio_code', 'performers', 'sequence', 'group', 'title'}
    for token in filename_tokens:
        if token.type not in labeled_types and token.value.strip():
            unlabeled_filename_tokens.add(token.value)

    # Calculate match stats
    total_filename_tokens = len(filename_tokens)
    matched_tokens = sum(1 for t in filename_tokens if t.type in labeled_types)
    match_rate = matched_tokens / total_filename_tokens if total_filename_tokens > 0 else 0.0

    match_stats = {
        # "path_tokens": 1 if path_token else 0,  # Disabled - not working on paths yet
        "filename_tokens": total_filename_tokens,
        "matched_tokens": matched_tokens,
        "match_rate": round(match_rate, 4)
    }

    return ParsedRow(
        input=filename,
        removed=removed_str,
        # path=path_token.value if path_token else None,  # Disabled - not working on paths yet
        filename_cleaned=result.cleaned,
        # path_pattern=path_pattern,  # Disabled - not working on paths yet
        filename_pattern=filename_pattern,
        studio=studio,
        title=title,
        performers=', '.join(performers) if performers else None,
        date=date,
        studio_code=studio_code,
        sequence=sequence,
        group=group,
        # unlabeled_path_tokens=unlabeled_path_tokens if unlabeled_path_tokens else None,  # Disabled - not working on paths yet
        unlabeled_filename_tokens=unlabeled_filename_tokens if unlabeled_filename_tokens else None,
        match_stats=match_stats
    )


def calculate_blind_metrics(rows: List[ParsedRow]) -> Dict[str, Any]:
    """
    Calculate coverage metrics for blind mode.

    Metrics include:
    - Field coverage (% of rows with each field populated)
    - Pattern histogram (top 20 patterns)
    - Anomalies (unmapped studios, unclassified tokens, etc.)
    """
    total_rows = len(rows)

    # Field coverage
    field_coverage = {
        'studio': sum(1 for r in rows if r.studio) / total_rows,
        'performers': sum(1 for r in rows if r.performers) / total_rows,
        'date': sum(1 for r in rows if r.date) / total_rows,
        'studio_code': sum(1 for r in rows if r.studio_code) / total_rows,
        'title': sum(1 for r in rows if r.title) / total_rows,
        'sequence': sum(1 for r in rows if r.sequence) / total_rows,
        'group': sum(1 for r in rows if r.group) / total_rows,
    }

    # Overall match rate
    avg_match_rate = sum(r.match_stats['match_rate'] for r in rows) / total_rows

    # Pattern histogram
    patterns = [r.filename_pattern for r in rows if r.filename_pattern]
    pattern_histogram = Counter(patterns).most_common(20)

    # Anomalies
    anomalies = {
        'rows_with_unlabeled_tokens': sum(1 for r in rows if r.unlabeled_filename_tokens),
        'total_unlabeled_tokens': sum(len(r.unlabeled_filename_tokens) for r in rows if r.unlabeled_filename_tokens),
    }

    return {
        'mode': 'blind',
        'total_rows': total_rows,
        'field_coverage': field_coverage,
        'avg_match_rate': round(avg_match_rate, 4),
        'pattern_histogram': [{'pattern': p, 'count': c} for p, c in pattern_histogram],
        'anomalies': anomalies,
        'timestamp': datetime.now().isoformat()
    }


def load_reference_data(filepath: Union[str, Path], sheet_name: str = "Reference") -> List[ParsedRow]:
    """
    Load reference data from Excel file.

    Expects the same 16-column schema as our output.
    """
    filepath = Path(filepath)
    wb = load_workbook(filepath, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Could not find '{sheet_name}' sheet in Excel file. Sheets present: {wb.sheetnames}")

        ws = wb[sheet_name]
        if ws is None:
            raise ValueError(f"Excel file has no '{sheet_name}' worksheet")

        # Get headers
        headers = [cell.value for cell in ws[1]]

        # Validate headers match our schema
        expected_headers = ParsedRow.get_headers()
        if headers != expected_headers:
            raise ValueError(f"Reference file headers don't match schema.\nExpected: {expected_headers}\nGot: {headers}")

        # Load rows
        reference_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Convert cell values to appropriate types
            def to_str(val) -> str:
                return str(val) if val is not None else ""

            def to_str_or_none(val) -> Optional[str]:
                return str(val) if val is not None else None

            # Parse the row into ParsedRow
            # Note: Adjusted indices because path-related columns are disabled
            parsed_row = ParsedRow(
                input=to_str(row[0]),
                removed=to_str(row[1]),
                # path=to_str_or_none(row[2]),  # Disabled - not working on paths yet
                filename_cleaned=to_str(row[2]),  # Was row[3], now row[2]
                # path_pattern=to_str_or_none(row[4]),  # Disabled - not working on paths yet
                filename_pattern=to_str(row[3]),  # Was row[5], now row[3]
                studio=to_str_or_none(row[4]),  # Was row[6], now row[4]
                title=to_str_or_none(row[5]),  # Was row[7], now row[5]
                performers=to_str_or_none(row[6]),  # Was row[8], now row[6]
                date=to_str_or_none(row[7]),  # Was row[9], now row[7]
                studio_code=to_str_or_none(row[8]),  # Was row[10], now row[8]
                sequence=json.loads(str(row[9])) if row[9] else None,  # Was row[11], now row[9]
                group=to_str_or_none(row[10]),  # Was row[12], now row[10]
                # unlabeled_path_tokens=ast.literal_eval(str(row[13])) if row[13] else None,  # Disabled - not working on paths yet
                unlabeled_filename_tokens=ast.literal_eval(str(row[11])) if row[11] else None,  # Was row[14], now row[11]
                match_stats=json.loads(str(row[12])) if row[12] else {}  # Was row[15], now row[12]
            )
            reference_rows.append(parsed_row)
    finally:
        wb.close()
    return reference_rows


def calculate_reference_metrics(rows: List[ParsedRow], reference_rows: List[ParsedRow], samples: int = 10) -> Dict[str, Any]:
    """
    Calculate comprehensive metrics for reference mode with user-defined metrics:

    1. Pattern Match Rate: % of files where pattern fields match exactly
    2. Metadata False Negative Rate: % of opportunities where reference has data but result is empty
    3. Metadata False Positive Rate: % of opportunities where reference is empty but result has data
    4. Metadata Accuracy Rate: % of opportunities where both have data and it matches exactly
    5. Parsed Perfect Rate: % of files where all 7 metadata fields match perfectly
    """
    total_rows = len(rows)
    metadata_fields = ['studio', 'performers', 'date', 'studio_code', 'title', 'sequence', 'group']

    if len(reference_rows) != total_rows:
        raise ValueError(f"Row count mismatch: {total_rows} parsed vs {len(reference_rows)} reference")

    # 1. Pattern Match Rate (only count opportunities where reference has a pattern)
    pattern_matches = 0
    pattern_opportunities = 0
    for i in range(total_rows):
        ref_pattern = reference_rows[i].filename_pattern
        parsed_pattern = rows[i].filename_pattern

        # Only count as an opportunity if reference has a pattern
        if ref_pattern is not None and ref_pattern != "":
            pattern_opportunities += 1
            if parsed_pattern == ref_pattern:
                pattern_matches += 1

    pattern_match_rate = (pattern_matches / pattern_opportunities * 100) if pattern_opportunities > 0 else 0.0

    # Initialize counters for metadata metrics
    false_negative_opportunities = 0
    false_negatives = 0
    false_positive_opportunities = 0
    false_positives = 0
    accuracy_opportunities = 0
    accurate_matches = 0
    files_perfectly_parsed = 0

    # Per-field tracking
    field_metrics = {}
    mismatches = []

    for field in metadata_fields:
        field_fn_opps = 0  # Opportunities for false negatives
        field_fns = 0      # False negatives
        field_fp_opps = 0  # Opportunities for false positives
        field_fps = 0      # False positives
        field_acc_opps = 0 # Opportunities for accuracy
        field_acc = 0      # Accurate matches

        field_mismatches = []

        for i in range(total_rows):
            parsed_val = getattr(rows[i], field)
            ref_val = getattr(reference_rows[i], field)

            # False Negative: Reference has data, result is empty
            if ref_val is not None and ref_val != "":
                field_fn_opps += 1
                false_negative_opportunities += 1
                if parsed_val is None or parsed_val == "":
                    field_fns += 1
                    false_negatives += 1
                    field_mismatches.append({
                        'input': rows[i].input,
                        'field': field,
                        'type': 'false_negative',
                        'parsed': parsed_val,
                        'expected': ref_val
                    })

            # False Positive: Reference is empty, result has data
            if ref_val is None or ref_val == "":
                field_fp_opps += 1
                false_positive_opportunities += 1
                if parsed_val is not None and parsed_val != "":
                    field_fps += 1
                    false_positives += 1
                    field_mismatches.append({
                        'input': rows[i].input,
                        'field': field,
                        'type': 'false_positive',
                        'parsed': parsed_val,
                        'expected': ref_val
                    })

            # Accuracy: Reference has data, check if result matches
            if ref_val is not None and ref_val != "":
                field_acc_opps += 1
                accuracy_opportunities += 1
                if parsed_val == ref_val:
                    field_acc += 1
                    accurate_matches += 1
                elif parsed_val is not None and parsed_val != "":
                    # Both have data but different - track as inaccuracy
                    field_mismatches.append({
                        'input': rows[i].input,
                        'field': field,
                        'type': 'incorrect',
                        'parsed': parsed_val,
                        'expected': ref_val
                    })

        # Calculate rates for this field
        field_fn_rate = (field_fns / field_fn_opps * 100) if field_fn_opps > 0 else 0.0
        field_fp_rate = (field_fps / field_fp_opps * 100) if field_fp_opps > 0 else 0.0
        field_acc_rate = (field_acc / field_acc_opps * 100) if field_acc_opps > 0 else 0.0

        field_metrics[field] = {
            'false_negative_rate': round(field_fn_rate, 2),
            'false_negative_count': field_fns,
            'false_negative_opportunities': field_fn_opps,
            'false_positive_rate': round(field_fp_rate, 2),
            'false_positive_count': field_fps,
            'false_positive_opportunities': field_fp_opps,
            'accuracy_rate': round(field_acc_rate, 2),
            'accurate_count': field_acc,
            'accuracy_opportunities': field_acc_opps,
        }

        # Add sample mismatches
        mismatches.extend(field_mismatches[:samples])

    # 5. Parsed Perfect Rate: Files where all 7 fields match exactly
    for i in range(total_rows):
        perfect = True
        for field in metadata_fields:
            if getattr(rows[i], field) != getattr(reference_rows[i], field):
                perfect = False
                break
        if perfect:
            files_perfectly_parsed += 1

    parsed_perfect_rate = (files_perfectly_parsed / total_rows * 100) if total_rows > 0 else 0.0

    # Calculate overall metadata rates
    metadata_false_negative_rate = (false_negatives / false_negative_opportunities * 100) if false_negative_opportunities > 0 else 0.0
    metadata_false_positive_rate = (false_positives / false_positive_opportunities * 100) if false_positive_opportunities > 0 else 0.0
    metadata_accuracy_rate = (accurate_matches / accuracy_opportunities * 100) if accuracy_opportunities > 0 else 0.0

    return {
        'mode': 'reference',
        'key_metrics': {
            'pattern_match_rate': round(pattern_match_rate, 2),
            'metadata_false_negative_rate': round(metadata_false_negative_rate, 2),
            'metadata_false_positive_rate': round(metadata_false_positive_rate, 2),
            'metadata_accuracy_rate': round(metadata_accuracy_rate, 2),
            'parsed_perfect_rate': round(parsed_perfect_rate, 2),
        },
        'summary': {
            'total_files': total_rows,
            'pattern_matches': pattern_matches,
            'pattern_opportunities': pattern_opportunities,
            'files_perfectly_parsed': files_perfectly_parsed,
            'false_negative_opportunities': false_negative_opportunities,
            'false_negatives': false_negatives,
            'false_positive_opportunities': false_positive_opportunities,
            'false_positives': false_positives,
            'accuracy_opportunities': accuracy_opportunities,
            'accurate_matches': accurate_matches,
        },
        'field_breakdown': field_metrics,
        'sample_mismatches': mismatches[:samples * 3],
        'timestamp': datetime.now().isoformat()
    }


def create_diff_rows(parsed_rows: List[ParsedRow], reference_rows: List[ParsedRow]) -> List[ParsedRow]:
    """
    Create diff rows showing differences between parsed and reference.

    For cells that differ, format as: {"expected": <ref>, "returned": <parsed>}
    For cells that match, keep the agreed value.
    """
    diff_rows = []

    for i in range(len(parsed_rows)):
        parsed = parsed_rows[i]
        reference = reference_rows[i]

        # Create diff row
        diff = ParsedRow(
            input=parsed.input,  # Always use parsed input
            removed=parsed.removed if parsed.removed == reference.removed else f'{{"expected": "{reference.removed}", "returned": "{parsed.removed}"}}',
            # path=parsed.path if parsed.path == reference.path else reference.path,  # Disabled - not working on paths yet
            filename_cleaned=parsed.filename_cleaned if parsed.filename_cleaned == reference.filename_cleaned else f'{{"expected": "{reference.filename_cleaned}", "returned": "{parsed.filename_cleaned}"}}',
            # path_pattern=parsed.path_pattern if parsed.path_pattern == reference.path_pattern else reference.path_pattern,  # Disabled - not working on paths yet
            filename_pattern=parsed.filename_pattern if parsed.filename_pattern == reference.filename_pattern else f'{{"expected": "{reference.filename_pattern}", "returned": "{parsed.filename_pattern}"}}',
            studio=parsed.studio if parsed.studio == reference.studio else f'{{"expected": "{reference.studio}", "returned": "{parsed.studio}"}}',
            title=parsed.title if parsed.title == reference.title else f'{{"expected": "{reference.title}", "returned": "{parsed.title}"}}',
            performers=parsed.performers if parsed.performers == reference.performers else f'{{"expected": "{reference.performers}", "returned": "{parsed.performers}"}}',
            date=parsed.date if parsed.date == reference.date else f'{{"expected": "{reference.date}", "returned": "{parsed.date}"}}',
            studio_code=parsed.studio_code if parsed.studio_code == reference.studio_code else f'{{"expected": "{reference.studio_code}", "returned": "{parsed.studio_code}"}}',
            sequence=parsed.sequence if parsed.sequence == reference.sequence else {"expected": reference.sequence, "returned": parsed.sequence},
            group=parsed.group if parsed.group == reference.group else f'{{"expected": "{reference.group}", "returned": "{parsed.group}"}}',
            # unlabeled_path_tokens=parsed.unlabeled_path_tokens if parsed.unlabeled_path_tokens == reference.unlabeled_path_tokens else reference.unlabeled_path_tokens,  # Disabled - not working on paths yet
            unlabeled_filename_tokens=parsed.unlabeled_filename_tokens if parsed.unlabeled_filename_tokens == reference.unlabeled_filename_tokens else reference.unlabeled_filename_tokens,
            match_stats=parsed.match_stats
        )

        diff_rows.append(diff)

    return diff_rows


def is_discrepancy_value(value: Any) -> bool:
    """Check if a cell value represents a discrepancy."""
    if value is None:
        return False
    value_str = str(value)
    # Check if the value contains the discrepancy format
    return '{"expected":' in value_str or '"expected":' in value_str


def write_excel_output(rows: List[ParsedRow], output_path: Union[str, Path], mode: str,
                       reference_rows: Optional[List[ParsedRow]] = None,
                       diff_rows: Optional[List[ParsedRow]] = None):
    """
    Write parsed results to Excel workbook.

    For blind mode: single sheet with results
    For reference mode: three sheets (Reference, Results, Diff)
    In reference mode, discrepancies in the Diff sheet are highlighted in yellow.
    """
    rows_values = [row.to_excel_row() for row in rows]
    headers = ParsedRow.get_headers()

    sheets: List[ExcelSheetData] = []
    if mode == 'blind':
        sheets.append(
            ExcelSheetData(
                name="Filename Parser Results",
                headers=headers,
                rows=rows_values,
            )
        )
    else:
        if reference_rows and diff_rows:
            sheets.append(
                ExcelSheetData(
                    name="Reference",
                    headers=headers,
                    rows=[row.to_excel_row() for row in reference_rows],
                )
            )
            sheets.append(
                ExcelSheetData(
                    name="Results",
                    headers=headers,
                    rows=rows_values,
                )
            )
            sheets.append(
                ExcelSheetData(
                    name="Diff",
                    headers=headers,
                    rows=[row.to_excel_row() for row in diff_rows],
                    highlight_discrepancies=True,
                    discrepancy_predicate=is_discrepancy_value,
                )
            )
        else:
            sheets.append(
                ExcelSheetData(
                    name="Results",
                    headers=headers,
                    rows=rows_values,
                )
            )

    write_excel_workbook(output_path, sheets)


def write_json_metrics(metrics: Dict[str, Any], output_path: Union[str, Path]):
    """Write metrics to JSON file."""
    output_path = Path(output_path)
    with output_path.open('w', encoding='utf-8') as f:
        json.dump(metrics, f, indent=2)


def main():
    """Main evaluation harness entry point."""
    args = parse_arguments()

    # Normalize paths and generate defaults
    input_path = Path(args.input)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    output_excel = Path(args.output_excel) if args.output_excel else Path("metrics") / f"{args.mode}-{timestamp}.xlsx"
    output_json = Path(args.output_json) if args.output_json else Path("metrics") / f"{args.mode}-{timestamp}.json"

    args.input = input_path
    args.output_excel = output_excel
    args.output_json = output_json

    print(f"=== Filename Parser Evaluation ({args.mode} mode) ===")
    print(f"Input: {args.input}")
    if not args.no_write:
        if not args.skip_excel:
            print(f"Excel output: {args.output_excel}")
        print(f"JSON output: {args.output_json}")

    # Read input filenames
    print("\nReading input...")
    reference_sheet_name = "Reference" if args.mode == 'reference' else None
    filenames = read_input_file(args.input, args.limit, sheet_name=reference_sheet_name)
    print(f"Found {len(filenames)} filenames to process")

    # Initialize parser
    parser = FilenameParser()

    # Parse all filenames
    print("\nParsing filenames...")
    rows = []
    for idx, filename in enumerate(filenames, 1):
        if idx % 100 == 0:
            print(f"  Processed {idx}/{len(filenames)}...")

        row = parse_filename(parser, filename)
        rows.append(row)

    print(f"Completed parsing {len(rows)} filenames")

    # Calculate metrics
    print(f"\nCalculating {args.mode} mode metrics...")
    reference_rows = None
    diff_rows = None

    if args.mode == 'blind':
        metrics = calculate_blind_metrics(rows)
    else:
        # Reference mode: load reference data and calculate diff
        print(f"Loading reference data from {args.input}...")
        reference_rows = load_reference_data(args.input, sheet_name="Reference")

        print(f"Comparing {len(rows)} parsed rows vs {len(reference_rows)} reference rows...")
        metrics = calculate_reference_metrics(rows, reference_rows, args.samples)

        # Create diff rows
        diff_rows = create_diff_rows(rows, reference_rows)

    # Print summary
    print("\n=== Metrics Summary ===")
    if args.mode == 'blind':
        print(f"Total rows: {metrics['total_rows']}")
        print(f"Average match rate: {metrics['avg_match_rate']:.2%}")
        print("\nField coverage:")
        for field, coverage in metrics['field_coverage'].items():
            print(f"  {field}: {coverage:.2%}")
        print(f"\nTop 5 patterns:")
        for item in metrics['pattern_histogram'][:5]:
            print(f"  {item['pattern']}: {item['count']} occurrences")
    else:
        # Reference mode summary
        key_metrics = metrics['key_metrics']
        summary = metrics['summary']
        field_breakdown = metrics['field_breakdown']

        print(f"\n{'='*60}")
        print(f"{'KEY METRICS':^60}")
        print(f"{'='*60}")
        print(f"Pattern Match Rate:              {key_metrics['pattern_match_rate']:>6.2f}%")
        print(f"Metadata False Negative Rate:    {key_metrics['metadata_false_negative_rate']:>6.2f}%")
        print(f"Metadata False Positive Rate:    {key_metrics['metadata_false_positive_rate']:>6.2f}%")
        print(f"Metadata Accuracy Rate:          {key_metrics['metadata_accuracy_rate']:>6.2f}%")
        print(f"Parsed Perfect Rate:             {key_metrics['parsed_perfect_rate']:>6.2f}%")

        print(f"\n{'='*60}")
        print(f"{'SUMMARY':^60}")
        print(f"{'='*60}")
        print(f"Total files:                     {summary['total_files']:>6}")
        print(f"Pattern matches:                 {summary['pattern_matches']:>6} / {summary['pattern_opportunities']:<6} opportunities")
        print(f"Files perfectly parsed:          {summary['files_perfectly_parsed']:>6}")
        print(f"\nFalse Negatives:                 {summary['false_negatives']:>6} / {summary['false_negative_opportunities']:<6} opportunities")
        print(f"False Positives:                 {summary['false_positives']:>6} / {summary['false_positive_opportunities']:<6} opportunities")
        print(f"Accurate matches:                {summary['accurate_matches']:>6} / {summary['accuracy_opportunities']:<6} opportunities")

        print(f"\n{'='*60}")
        print(f"{'FIELD BREAKDOWN':^60}")
        print(f"{'='*60}")
        for field, field_data in field_breakdown.items():
            print(f"\n{field.upper()}")
            print(f"  False Negative Rate: {field_data['false_negative_rate']:>6.2f}%  ({field_data['false_negative_count']}/{field_data['false_negative_opportunities']} opportunities)")
            print(f"  False Positive Rate: {field_data['false_positive_rate']:>6.2f}%  ({field_data['false_positive_count']}/{field_data['false_positive_opportunities']} opportunities)")
            print(f"  Accuracy Rate:       {field_data['accuracy_rate']:>6.2f}%  ({field_data['accurate_count']}/{field_data['accuracy_opportunities']} opportunities)")

        if metrics.get('sample_mismatches'):
            print(f"\n{'='*60}")
            print(f"{'SAMPLE MISMATCHES':^60}")
            print(f"{'='*60}")
            print(f"Showing up to {args.samples} per field:")
            for mismatch in metrics['sample_mismatches'][:10]:
                mismatch_type = mismatch.get('type', 'unknown')
                print(f"\n  [{mismatch_type.upper()}] {mismatch['field']}")
                print(f"  File: {mismatch['input'][:70]}...")
                print(f"    Expected: {mismatch['expected']}")
                print(f"    Parsed:   {mismatch['parsed']}")

    # Write outputs
    if not args.no_write:
        if not args.skip_excel:
            print(f"\nWriting Excel output to {args.output_excel}...")
            write_excel_output(rows, args.output_excel, args.mode,
                             reference_rows=reference_rows,
                             diff_rows=diff_rows)

        print(f"Writing JSON metrics to {args.output_json}...")
        write_json_metrics(metrics, args.output_json)

        print("\n✓ Evaluation complete!")
    else:
        print("\n✓ Dry-run complete (no files written)")


if __name__ == '__main__':
    main()
